import asyncio
import asyncssh
from collections import namedtuple
from typing import AsyncGenerator
from zipfile import BadZipFile

import openpyxl
from telegram import Update
from transliterate import slugify

from Clases.ApiMarketplaces.Ali.AliApiAsync import AliApiAsync
from Clases.ApiMarketplaces.Ozon.OzonApiAsync import OzonApiAsync
from Clases.ApiMarketplaces.Ozon.OzonProdResponse import OzonProdResponse
from Clases.ApiMarketplaces.Ozon.Warehouse import Warehouse
from Clases.ApiMarketplaces.Vk.VKProduct import VkProduct
from Clases.ApiMarketplaces.Vk.VkApiAsync import VkApiAsync
from Clases.ApiMarketplaces.Ya.YAapiAsync import YAapiAsync
from Clases.BifitApi.Good import Good
from Clases.BifitApi.Request import Request
from logger import logger
from methods.sync_methods import get_vk_skus_id_dict


async def send_to_yandex_async_v2(ya_token: str,
                                  ya_campaign_id: int,
                                  ya_warehouse_id: int,
                                  ya_goods_set: set[Good]) -> dict:
    """
    Отправляет данные об остатках товаров на складе в Яндекс.Маркет.

    :param ya_token: токен для доступа к API Яндекс.Маркета
    :param ya_campaign_id: идентификатор кампании в Яндекс.Маркете
    :param ya_warehouse_id: идентификатор склада в Яндекс.Маркете
    :param ya_goods_set: множество товаров Яндекс-маркета
    :return: словарь с ответом сервера, в случае с ошибки или пустой словарь, если ошибок нет
    """
    logger.debug(f'send_to_yandex_async started')
    ya_api = YAapiAsync(ya_token, ya_campaign_id, ya_warehouse_id, goods_set=ya_goods_set)
    ya_send_remains_response = await ya_api.send_remains_async()
    if 'errors' in ya_send_remains_response:
        logger.error(f'ошибка на этапе отправки товаров в Яндекс - {ya_send_remains_response}')
        return ya_send_remains_response
    logger.debug(f'send_to_yandex_async finished smoothly')
    return {}


async def send_to_vk_async_v2(vk_token: str,
                              vk_owner_id: int,
                              vk_api_ver: float,
                              vk_goods_set: set[Good]) -> dict:
    logger.debug('send_to_vk_async started')
    vk_api = VkApiAsync(vk_token, vk_owner_id, vk_api_ver)
    vk_products_items = await vk_api.get_all_products_async()
    if 'error' in vk_products_items:
        logger.error(f'ошибка на этапе запроса товаров в ВК - {vk_products_items[1]}')
        return {'error': f'Ошибка в ответе сервера на запрос товаров ВК - {vk_products_items[1]}'}
    try:
        vk_products = {VkProduct(item) for item in vk_products_items}
    except KeyError as e:
        logger.error(f'Ошибка формирования списка товаров - {e}')
        logger.debug('send_to_vk_async finished with exception')
        return {'error': f'Ошибка формирования списка товаров ВК - {str(e)}'}

    logger.debug('send_to_vk_async finished smoothly')
    return await vk_api.send_remains_async_v2(vk_goods_set, vk_products)


async def send_to_ozon_async_v2(ozon_admin_key: str, ozon_client_id: str, ozon_goods_set: set[Good]) -> dict:
    ozon_session = OzonApiAsync(ozon_admin_key, ozon_client_id)

    ozon_products_request = await ozon_session.get_all_products_async_v3()
    logger.debug(f' ozon_products_request - {ozon_products_request}')
    if 'error' in ozon_products_request:
        logger.error(f'send_to_ozon_async finished with error - {ozon_products_request}')
        return ozon_products_request
    try:
        ozon_products_response = OzonProdResponse(ozon_products_request)
    except KeyError as e:
        logger.error(f'Ошибка в ответе сервера на запрос товаров Озон - {e}')
        return {'error': f'Ошибка в ответе сервера на запрос товаров Озон - {str(e)}'}
    ozon_products_dict = ozon_products_response.get_skus_id_dict()
    logger.debug(f' ozon_products_dict - {ozon_products_dict}')
    ozon_warehouses_response = await ozon_session.get_warehouses_async()
    logger.debug(f' ozon_warehouses_response - {ozon_warehouses_response}')
    warehouses_result_list = ozon_warehouses_response.get('result')
    if not warehouses_result_list:
        logger.error(f'Ошибка в ответе сервера на запрос списка складов Озон - {ozon_warehouses_response}')
        return {'error': f'Ошибка в ответе сервера на запрос списка складов - {ozon_warehouses_response}'}

    ozon_warehouses = [Warehouse(w) for w in warehouses_result_list if w.get('status') != "disabled"]
    return await ozon_session.send_remains_async_v2(ozon_products_dict, ozon_goods_set, ozon_warehouses)


async def send_to_ozon_stores_v2(
        ozon_keys: dict[str, str],
        ozon_goods_set: set[Good]) -> list[dict] | None:
    """
    Отправляет данные в магазины Ozon параллельно и собирает возможные ошибки.

    :param ozon_keys: Словарь с ключами доступа к магазинам Ozon (id: key).
    :param ozon_goods_set: Множество с товарами для отправки.
    :return: Список ошибок, если они есть, либо None.
    """
    logger.debug(f'Начал send_to_ozon_stores для %d магазинов.', len(ozon_keys))

    errors_coroutines = [
        send_to_ozon_async_v2(oz_key, oz_id, ozon_goods_set)
        for oz_id, oz_key in ozon_keys.items()
    ]
    try:
        # Запускаем все корутины параллельно и собираем результаты
        errors = await asyncio.gather(*errors_coroutines, return_exceptions=True)
    except Exception as e:
        logger.error("Произошла ошибка при выполнении send_to_ozon_stores: %s", e)
        return [{"error": "Critical error", "details": str(e)}]

    filtered_errors = [error for error in errors if isinstance(error, dict) or isinstance(error, Exception)]
    if any(filtered_errors):
        logger.warning(f'Обнаружены ошибки при отправке данных в Ozon: {filtered_errors}')
        return filtered_errors
    logger.debug('Закончил send_to_ozon_stores без ошибок.')
    return None


async def read_xlsx_async(file_path_name: str, update: Update) -> AsyncGenerator:
    """Генератор, читает файл построчно"""
    logger.debug('начал read_xlsx')
    required_fields = {'barcode', 'selling_price', 'purchase_price'}

    # Открываем файл
    try:
        workbook = openpyxl.load_workbook(file_path_name)
    except FileNotFoundError:
        logger.error('не нашел xlsx')
        await update.message.reply_text('не нашел xlsx')
        return
    except BadZipFile:
        logger.error('этот файл не xlsx')
        await update.message.reply_text('этот файл не xlsx')
        return
    else:
        logger.debug('успешно открыл xlsx')
        await update.message.reply_text('успешно открыл xlsx')
        sheet = workbook.active

    # Получаем заголовки колонок
    headers = [cell.value for cell in sheet[1]]

    # Проверяем наличие необходимых заголовков
    missing_fields = required_fields - set(headers)

    if missing_fields:
        logger.error('Отсутствуют обязательные поля в файле - %s', missing_fields)
        await update.message.reply_text(f'Отсутствуют обязательные поля в файле - {missing_fields}')
        return

    # Создаем namedtuple динамически
    ExcelGood = namedtuple('Good', ['barcode', 'selling_price', 'purchase_price'])

    # Получаем индексы необходимых колонок
    field_indices = {field: headers.index(field) for field in required_fields}

    # Пропускаем заголовок
    total_rows = sheet.max_row - 1
    if total_rows == 0:
        logger.error("Файл не содержит данных для обработки.")
        await update.message.reply_text("Файл не содержит данных для обработки.")
        return

    progress_message = await update.message.reply_text("Начинаю обработку файла...")
    update_frequency = max(1, total_rows // 10)

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        # Извлекаем значения для каждого необходимого поля
        barcode = row[field_indices['barcode']]
        selling_price = row[field_indices['selling_price']]
        purchase_price = row[field_indices['purchase_price']]
        if barcode and selling_price is not None:
            # Создаем namedtuple для каждой строки
            excel_good = ExcelGood(str(barcode), selling_price, purchase_price)
            logger.debug('excel_good= %s', excel_good)
            yield excel_good

        # Обновление прогресса каждые 10%
        if i % update_frequency == 0:
            progress = i / total_rows * 100
            await progress_message.edit_text(f"Обработка: {progress:.0f}%")
            logger.info(f"Обработано {i} строк из {total_rows}, прогресс: {progress:.0f}%")

    await progress_message.edit_text("Обработка файла завершена!")


async def get_barcodes_from_xlsx_async(file_path_name: str, update: Update) -> dict[str, tuple[float, float]]:
    logger.debug('начал get_barcodes_from_xlsx')
    excel_goods = {}
    async for code, selling_price, purchase_price in read_xlsx_async(file_path_name, update):
        excel_goods[code] = (selling_price, purchase_price)
    return excel_goods


async def send_to_ali_async_v2(ali_token: str, ali_goods_set: set[Good]) -> dict:
    ali_api = AliApiAsync.from_goods_set(ali_token, ali_goods_set)
    err = await ali_api.fill_get_products_to_send()
    if err:
        return err
    return await ali_api.send_remains_async()


async def get_pic_url(pic_name: str, vendor_name: str, to_bot: bool = False) -> str:
    """Формирует ссылку на изображение товара"""
    logger.debug('get_pic_url started')
    my_site_url = 'https://pronogti.store'
    vendor_name = slugify(vendor_name, 'uk')
    pic_name = slugify(pic_name, 'uk')
    pic_url = f'{my_site_url}/images/{vendor_name}/{pic_name}.jpg'
    if to_bot:
        return pic_url
    logger.debug('сформировал ссылку на картинку товара\n'
                 f'{pic_url=}. отправляю запрос')
    pic_request = Request(url=pic_url)
    pic_response = await pic_request.send_get_async()
    if 'error' in pic_response:
        logger.error('ERROR ошибка при получении изображения\n'
                     f'ставлю заглушку {my_site_url}/images/no-image.jpg\n'
                     'get_pic_url finished with exception')
        return f'{my_site_url}/images/no-image.jpg'
    logger.debug('ОК нашел картинку на сервере.\n'
                 'get_pic_url finished smoothly')
    return pic_url

async def upload_file_async(local_path: str, remote_path: str, hostname: str, username: str):
    """Загружает файл на удалённый сервер по SSH с помощью ключа"""
    logger.debug('upload_file_async started')

    try:
        async with asyncssh.connect(
                hostname,
                username=username,
                known_hosts=None
        ) as conn:
            await asyncssh.scp(local_path, (conn, remote_path))
            logger.info('Файл %s успешно загружен', local_path)
            logger.debug('upload_file_async finished smoothly')
    except (asyncssh.Error, OSError) as e:
        logger.error("❌ Ошибка при загрузке файла: %s", e)
        logger.debug('upload_file_async finished with error')